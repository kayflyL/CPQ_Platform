"""Generate SQL script from Excel for importing opportunities."""
import sys
import json
import openpyxl
from pathlib import Path


def generate_sql(excel_path: str, output_sql_path: str):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print(f"Sheet: {ws.title}, Rows: {ws.max_row - 1}")

    inserted = 0
    skipped = 0

    with open(output_sql_path, "w", encoding="utf-8") as f:
        f.write("BEGIN;\n\n")

        index = 0
        for r in range(2, ws.max_row + 1):
            month = ws.cell(r, 1).value
            day = ws.cell(r, 2).value
            sales_person = ws.cell(r, 3).value
            fae = ws.cell(r, 4).value
            opportunity_name = ws.cell(r, 5).value
            chassis_form = ws.cell(r, 6).value
            platform_type = ws.cell(r, 7).value
            config_count = ws.cell(r, 8).value
            purchase_qty = ws.cell(r, 9).value
            feedback = ws.cell(r, 10).value
            feedback_content = ws.cell(r, 11).value

            if not opportunity_name:
                skipped += 1
                continue

            index += 1

            sales_person = sales_person or ""
            fae = fae or ""
            chassis_form = chassis_form or ""
            platform_type = platform_type or ""
            config_count = config_count or 0
            purchase_qty = purchase_qty or 0

            created_at = f"2026-{month:02d}-{day:02d} 00:00:00"
            quotation_date = f"2026-{month:02d}-{day:02d}"

            opp_id = f"OPP-2026{month:02d}{day:02d}-{index:03d}"
            quo_id = f"Q-{opp_id}"

            extra = {}
            if feedback:
                extra["feedback"] = str(feedback)
            if feedback_content:
                extra["feedback_content"] = str(feedback_content)
            extra_json = json.dumps(extra, ensure_ascii=False) if extra else ""

            def esc(s):
                return str(s).replace("'", "''")

            opp_name = esc(opportunity_name)

            # Insert opportunity
            f.write(f"-- Row {r}: {opportunity_name[:40]}\n")
            f.write(f"INSERT INTO opportunities.opportunities (\n")
            f.write(f"    opportunity_id, folder_name, opportunity_name, customer_name,\n")
            f.write(f"    sales_person, fae, quotation_person, platform_type, chassis_form,\n")
            f.write(f"    purchase_qty, created_at, updated_at, status, extra_fields, tenant_id\n")
            f.write(f") VALUES (\n")
            f.write(f"    '{opp_id}', '', '{opp_name}', '',\n")
            f.write(f"    '{esc(sales_person)}', '{esc(fae)}', '', '{esc(platform_type)}', '{esc(chassis_form)}',\n")
            f.write(f"    {purchase_qty}, '{created_at}', '{created_at}', 'archived',\n")
            extra_val = 'NULL' if not extra else f"'{esc(extra_json)}'"
            f.write(f"    {extra_val}, 'default'\n")
            f.write(f");\n\n")

            # Insert placeholder quotation
            f.write(f"INSERT INTO opportunities.quotations (\n")
            f.write(f"    quotation_id, opportunity_id, version, quotation_name, file_path,\n")
            f.write(f"    l6_price, total_qty, config_count, created_at, updated_at, status,\n")
            f.write(f"    config_quantities, config_descriptions, config_server_models,\n")
            f.write(f"    config_warranty_info, total_price, profit_margin, extra_fields,\n")
            f.write(f"    tenant_id, quotation_date, is_primary\n")
            f.write(f") VALUES (\n")
            f.write(f"    '{quo_id}', '{opp_id}', 'v1', '{opp_name}', NULL,\n")
            f.write(f"    0, {purchase_qty}, {config_count}, '{created_at}', '{created_at}', 'active',\n")
            f.write(f"    NULL, NULL, NULL, NULL, 0, 0, NULL,\n")
            f.write(f"    'default', '{quotation_date}', true\n")
            f.write(f");\n\n")

            inserted += 1

        f.write("COMMIT;\n")

    print(f"Generated: {inserted} opportunities, {skipped} skipped")
    print(f"SQL written to: {output_sql_path}")


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else str(
        Path.home() / ".hermes" / "desktop-attachments" / "项目List-2.xlsx"
    )
    output = excel_path.replace(".xlsx", ".sql")
    generate_sql(excel_path, output)
