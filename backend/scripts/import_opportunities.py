"""Import historical opportunities from Excel file into the database."""
import sys
import json
from pathlib import Path
from datetime import datetime

# Add backend to path
backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

from sqlalchemy import create_engine, text
from app.core.config import get_settings
settings = get_settings()


def generate_opportunity_id(month: int, day: int, index: int) -> str:
    """Generate unique opportunity ID."""
    return f"OPP-2026{month:02d}{day:02d}-{index:03d}"


def normalize_platform(platform_type: str | None) -> str | None:
    """Normalize platform names: 兆芯→Polaris, INTEL→Intel."""
    if not platform_type:
        return None
    p = str(platform_type).strip()
    if p == '兆芯':
        return 'Polaris'
    if p.upper() == 'INTEL':
        return 'Intel'
    if p == 'INTEL&Orion':
        # Mixed platform — keep as-is, will be handled by caller
        return p
    return p


def clean_sales_person(sales: str | None, opportunity_name: str | None) -> tuple[str | None, str]:
    """Fix Row 100-style rows where sales column contains opportunity name."""
    if not sales:
        return None, (opportunity_name or '')
    sales = str(sales).strip()
    name = str(opportunity_name or '').strip()
    if len(sales) > 10 and any(kw in sales for kw in ['4U', '2U', '卡', '机头', '配置']):
        # Extract text before first spec keyword (4U/2U/卡/etc.)
        import re
        # Find first occurrence of any spec keyword
        pos = len(sales)
        for kw in ['4U', '2U', '5U', '4.5U', '8U', '卡', '机头', '配置']:
            idx = sales.find(kw)
            if 0 <= idx < pos:
                pos = idx
        extracted = sales[:pos].rstrip('-')
        return extracted if extracted else None, name if name else sales
    return sales, name if name else sales


def import_opportunities(excel_path: str):
    """Import rows from Excel into opportunities + quotation placeholder records."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print(f"Sheet: {ws.title}, Rows: {ws.max_row - 1}")

    engine = create_engine(settings.DATABASE_URL)

    inserted = 0
    skipped = 0
    errors = 0
    mixed_platform_rows = []

    with engine.begin() as conn:
        index = 0
        for r in range(2, ws.max_row + 1):
            month = ws.cell(r, 1).value
            day = ws.cell(r, 2).value
            sales_person = ws.cell(r, 3).value
            fae = ws.cell(r, 4).value
            opportunity_name = ws.cell(r, 5).value
            chassis_form = ws.cell(r, 6).value
            platform_type = ws.cell(r, 7).value
            config_count = ws.cell(r, 8).value
            purchase_qty = ws.cell(r, 9).value
            feedback = ws.cell(r, 10).value
            feedback_content = ws.cell(r, 11).value

            # Skip rows with no opportunity name
            if not opportunity_name:
                skipped += 1
                continue

            index += 1

            # Fix messy sales person (Row 100 case)
            sales_person, opportunity_name = clean_sales_person(sales_person, opportunity_name)

            # Normalize platform
            if platform_type and 'INTEL&Orion' in str(platform_type):
                mixed_platform_rows.append(r)
            platform_type = normalize_platform(platform_type)

            # Clean None values
            sales_person = sales_person or None
            fae = fae or None
            chassis_form = chassis_form or None
            config_count = config_count or 0
            purchase_qty = purchase_qty or 0

            # Date string
            created_at = f"2026-{month:02d}-{day:02d} 00:00:00"

            opportunity_id = generate_opportunity_id(month, day, index)

            # Build extra_fields JSON for feedback columns
            extra = {}
            if feedback:
                extra["feedback"] = str(feedback)
            if feedback_content:
                extra["feedback_content"] = str(feedback_content)
            extra_fields = json.dumps(extra, ensure_ascii=False) if extra else None

            # Insert opportunity (status=archived)
            conn.execute(
                text("""
                    INSERT INTO opportunities.opportunities (
                        opportunity_id, folder_name, opportunity_name, customer_name,
                        sales_person, fae, quotation_person, platform_type, chassis_form,
                        purchase_qty, created_at, updated_at, status, extra_fields, tenant_id
                    ) VALUES (
                        :opportunity_id, :folder_name, :opportunity_name, '',
                        :sales_person, :fae, '', :platform_type, :chassis_form,
                        :purchase_qty, :created_at, :updated_at, 'archived', :extra_fields, 'default'
                    )
                """),
                {
                    "opportunity_id": opportunity_id,
                    "folder_name": "",
                    "opportunity_name": str(opportunity_name),
                    "sales_person": sales_person,
                    "fae": fae,
                    "platform_type": platform_type,
                    "chassis_form": chassis_form,
                    "purchase_qty": int(purchase_qty),
                    "created_at": created_at,
                    "updated_at": created_at,
                    "extra_fields": extra_fields,
                },
            )

            # Insert placeholder quotation (is_primary=true)
            quotation_id = f"Q-{opportunity_id}"
            conn.execute(
                text("""
                    INSERT INTO opportunities.quotations (
                        quotation_id, opportunity_id, version, quotation_name, file_path,
                        l6_price, total_qty, config_count, created_at, updated_at, status,
                        config_quantities, config_descriptions, config_server_models,
                        config_warranty_info, total_price, profit_margin, extra_fields,
                        tenant_id, quotation_date, is_primary
                    ) VALUES (
                        :quotation_id, :opportunity_id, 'v1', :quotation_name, NULL,
                        0, :total_qty, :config_count, :created_at, :updated_at, 'active',
                        NULL, NULL, NULL, NULL, 0, 0, NULL,
                        'default', :quotation_date, true
                    )
                """),
                {
                    "quotation_id": quotation_id,
                    "opportunity_id": opportunity_id,
                    "quotation_name": str(opportunity_name),
                    "total_qty": int(purchase_qty),
                    "config_count": int(config_count),
                    "created_at": created_at,
                    "updated_at": created_at,
                    "quotation_date": f"2026-{month:02d}-{day:02d}",
                },
            )

            inserted += 1

    if mixed_platform_rows:
        print(f"\n⚠ Mixed platform rows (INTEL&Orion): {mixed_platform_rows}")

    print(f"\nDone: {inserted} imported, {skipped} skipped, {errors} errors")
    return inserted


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else str(
        Path.home() / ".hermes" / "desktop-attachments" / "项目List-3.xlsx"
    )
    print(f"Importing from: {excel_path}")
    count = import_opportunities(excel_path)
    print(f"Successfully imported {count} opportunities (with placeholder quotations).")
