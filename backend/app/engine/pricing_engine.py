"""
Pricing Engine — pure business logic layer.

Replaces the DB-coupled parts of legacy data_processor.py.
All data access goes through injected Repository instances.
No sqlite3, no direct DB connections.
"""

import re
import os
import ast
import json
import operator
import io
import pandas as pd
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Optional

from app.core.config import get_settings

from app.repository.kp_repo import KPRepository
from app.repository.l6_repo import L6Repository
from app.repository.opportunity_repo import OpportunityRepository
from app.repository.rules_repo import RulesRepository
from app.repository.univer_template_repo import UniverTemplateRepo
from app.engine.excel_parser import ExcelParser


# === Safe arithmetic evaluator (replaces eval() for Excel formulas) ===
_SAFE_BIN_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}
_SAFE_UNARY_OPS = {
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def _safe_eval_math(expr: str) -> float:
    """Safely evaluate a simple arithmetic expression (numbers + - * / parentheses).

    Unlike eval(), this does NOT support ** (exponentiation), so it cannot be
    abused for DoS via expressions like 9**9**9.
    """
    tree = ast.parse(expr, mode='eval')

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in _SAFE_BIN_OPS:
            return _SAFE_BIN_OPS[type(node.op)](_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _SAFE_UNARY_OPS:
            return _SAFE_UNARY_OPS[type(node.op)](_eval(node.operand))
        raise ValueError(f"Unsafe or unsupported expression: {ast.dump(node)}")

    return _eval(tree)


# === V49 Excel export style constants (unchanged from legacy) ===
_F_22_LISU = Font(name='隶书', size=22, bold=False, color='000000')
_F_11_SONG = Font(name='宋体', size=11, bold=False, color='000000')
_F_11_TNR = Font(name='Times New Roman', size=11, bold=False, color='000000')
_F_11_DENG_B = Font(name='等线', size=11, bold=True, color='000000')
_F_11_TNR_C = Font(name='Times New Roman', size=11, bold=False, color='000000')
_F_11_SONG_11 = Font(name='宋体', size=11, bold=False, color='000000')
_F_11_TNR_B = Font(name='Times New Roman', size=11, bold=True, color='000000')
_F_11_EQ = Font(name='等线', size=11, bold=False, color='000000')

_NO_FILL = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
_THIN_B = Border(
    left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
)

def _s(cell, font=None, fill=None, align=None, border=True):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = _THIN_B


# Data paths (from config)
_settings = get_settings()
DATA_DIR = Path(_settings.DATA_PATH)
CONFIG_PATH = DATA_DIR / "config.json"


class PricingEngine:
    """
    Business logic engine.
    Receives Repositories via constructor — never creates DB connections directly.
    """

    def __init__(self, kp_repo: KPRepository, l6_repo: L6Repository,
                 opportunity_repo: OpportunityRepository, rules_repo: RulesRepository = None,
                 quotation_repo=None):
        self.kp_repo = kp_repo
        self.l6_repo = l6_repo
        self.opportunity_repo = opportunity_repo
        self.rules_repo = rules_repo
        self._quotation_repo = quotation_repo
        
        # Initialize ExcelParser if rules_repo is available
        self._excel_parser = ExcelParser(rules_repo) if rules_repo else None

        # Load rules from DB (with hardcoded fallbacks)
        self._load_rules()

    def _get_quotation_repo(self):
        """Lazy-init QuotationRepository (avoids per-call instantiation)."""
        if self._quotation_repo is None:
            from app.repository.quotation_repo import QuotationRepository
            self._quotation_repo = QuotationRepository()
        return self._quotation_repo
    
    def _load_rules(self):
        """Load configurable rules from rules.db, with hardcoded fallbacks."""
        # Default hardcoded fallbacks for L6/KP region configs
        self._l6_region_config = {
            "region_start_keywords": "L6",
            "field_mapping": {"catalogue": "D", "description": "E", "quantity": "F"},
            "region_end_keywords": "Keyparts,KP"
        }
        self._kp_region_config = {
            "region_start_keywords": "Keyparts,KP",
            "field_mapping": {"catalogue": "D", "model": "E", "quantity": "F", "price": "G"},
            "region_end_keywords": "Warranty,Total"
        }
        self._kp_cat_map = {
            'cpu': 'CPU', 'processor': 'CPU',
            'memory': 'Memory', 'ram': 'Memory',
            'hdd': 'HDD/SSD', 'ssd': 'HDD/SSD',
            'raid': 'Raid card',
            'network': 'NIC', 'nic': 'NIC',
            'gpu': 'GPU',
            'power': 'Power', 'psu': 'Power',
            'fan': 'Fan',
            'heatsink': 'Heatsink', 'cooler': 'Heatsink',
            'cable': 'Cable', 'wire': 'Cable',
            'rail': 'Rail'
        }
        self._price_diff_threshold = 0.01
        
        if not self.rules_repo:
            return
        
        # Override from DB
        try:
            l6_config = self.rules_repo.get_l6_region_config()
            if l6_config:
                self._l6_region_config = l6_config
            
            kp_config = self.rules_repo.get_kp_region_config()
            if kp_config:
                self._kp_region_config = kp_config
            
            kp_mappings = self.rules_repo.get_kp_category_mappings()
            if kp_mappings:
                self._kp_cat_map = {m['keyword']: m['category'] for m in kp_mappings}
                # Store raw mappings for heatmap preview
                self._kp_cat_mappings_raw = kp_mappings
        except Exception as e:
            print(f"⚠️ Failed to load rules from DB, using defaults: {e}")

    # ==================== 1. Excel Parsing (pure algorithm) ====================

    def parse_file(self, sheet_dict: dict) -> tuple:
        """Parse uploaded Excel into configs + first_meta.
        
        Uses ExcelParser if available (new rule-driven approach),
        falls back to legacy _extract_meta + _parse_items otherwise.
        """
        configs = {}
        first_meta = None
        
        for sheet_name, df in sheet_dict.items():
            if '原始需求' in sheet_name or 'Reference' in sheet_name or df.empty:
                continue
            
            # Try new ExcelParser first
            if self._excel_parser:
                try:
                    parse_result = self._excel_parser.parse(df, return_trace=False)
                    meta = self._convert_parser_meta(parse_result["static_fields"])
                    items = self._convert_parser_items(parse_result["dynamic_regions"])
                except Exception as e:
                    print(f"⚠️ ExcelParser failed for sheet '{sheet_name}': {e}, falling back to legacy")
                    meta = self._extract_meta(df)
                    items = self._parse_items(df)
            else:
                # Legacy approach
                meta = self._extract_meta(df)
                items = self._parse_items(df)
            
            if items.empty:
                continue
            
            configs[sheet_name] = {'meta': meta, 'items': items}
            if first_meta is None:
                first_meta = meta
        
        return configs, first_meta
    
    def _convert_parser_meta(self, static_fields: dict) -> dict:
        """Convert ExcelParser static_fields to legacy meta format."""
        meta = {}
        
        # Map field keys to legacy meta keys
        field_mapping = {
            "project_name": "opportunity_name",
            "model_name": "model_name",
            "fae": "fae",
            "quotation_date": "date",
            "description": "l6_desc"
        }
        
        for parser_key, meta_key in field_mapping.items():
            if parser_key in static_fields:
                value = static_fields[parser_key]["value"]
                meta[meta_key] = value
                
                # Special handling for model_name (extract qty from parentheses)
                if parser_key == "model_name" and value:
                    m = re.search(r'\((\d+)', value)
                    if m:
                        meta['model_qty'] = m.group(1)
                        meta['model_name'] = value.split('(')[0].strip()
        
        return meta
    
    def _convert_parser_items(self, dynamic_regions: dict) -> pd.DataFrame:
        """Convert ExcelParser dynamic_regions to legacy items DataFrame format."""
        items = []
        
        # L6 region
        if "L6" in dynamic_regions:
            for item in dynamic_regions["L6"]:
                catalogue = item.get("l6_chassis", "")
                description = item.get("spec", "")
                qty = 1
                if "qty" in item:
                    try:
                        qty = int(float(item["qty"]))
                    except:
                        qty = 1
                
                if not catalogue or catalogue.lower() in ['nan', 'none', '', 'catalogue']:
                    continue
                
                items.append({
                    'category': 'L6',
                    'part_name': catalogue,
                    'spec': description,
                    'qty': qty,
                    'currency': 'RMB'
                })
        
        # KP region
        if "KP" in dynamic_regions:
            for item in dynamic_regions["KP"]:
                catalogue = item.get("kp_category", "")
                model = item.get("kp_model", "")
                qty = 1
                if "qty" in item:
                    try:
                        qty = int(float(item["qty"]))
                    except:
                        qty = 1
                
                price = None
                if "kp_price" in item:
                    try:
                        price = float(item["kp_price"])
                    except:
                        pass
                
                if not catalogue or catalogue.lower() in ['nan', 'none', '', 'catalogue']:
                    continue
                
                # Determine if USD
                is_usd = False
                if 'cpu' in catalogue.lower() or 'processor' in catalogue.lower():
                    if 'usd' in model.lower() or '$' in model:
                        is_usd = True
                
                items.append({
                    'category': 'Key Parts',
                    'part_name': catalogue,
                    'spec': model,
                    'qty': qty,
                    'currency': 'USD' if is_usd else 'RMB'
                })
        
        # Warranty region
        if "Warranty" in dynamic_regions:
            for item in dynamic_regions["Warranty"]:
                warranty_type = item.get("part_name", "")
                description = item.get("description", "")
                
                if not description or description.lower() in ['nan', 'none', '']:
                    continue
                
                # Extract warranty years — let user fill in manually
                years = None
                
                items.append({
                    'category': 'Warranty',
                    'part_name': warranty_type,
                    'spec': description,
                    'qty': 1,
                    'currency': 'RMB',
                    'warranty_years': years
                })
        
        return pd.DataFrame(items) if items else pd.DataFrame()

    def preview_parse(self, df: pd.DataFrame, max_row: int = 15, max_col: int = 15, kp_mappings: list = None) -> dict:
        """Parse a single sheet for heatmap preview. Returns grid data + cell marks + meta."""
        meta, cell_marks = self._extract_meta(df, return_cell_marks=True)
        
        # Mark L6 region on the heatmap
        l6_cfg = self._l6_region_config
        l6_start = self._find_region_row(df, l6_cfg.get('region_start_keywords', 'L6'))
        l6_end = self._find_region_row(df, l6_cfg.get('region_end_keywords', 'Keyparts,KP'), start_row=l6_start + 1 if l6_start >= 0 else 0)
        
        if l6_start >= 0:
            end_row = l6_end if l6_end > l6_start else len(df)
            # Mark start keyword row
            for c in range(min(10, df.shape[1])):
                cell_val = str(df.iloc[l6_start, c]).strip() if pd.notna(df.iloc[l6_start, c]) else ''
                if cell_val:
                    cell_marks.append({'row': l6_start, 'col': c, 'value': cell_val, 'type': 'l6_region', 'target': 'L6 Start'})
            # Mark data rows
            l6_field_map = l6_cfg.get('field_mapping', {})
            if isinstance(l6_field_map, str):
                try: l6_field_map = json.loads(l6_field_map)
                except: l6_field_map = {}
            for r in range(l6_start + 1, end_row):
                for field_name, col_letter in l6_field_map.items():
                    col_idx = self._col_letter_to_index(col_letter)
                    if col_idx < df.shape[1]:
                        cell_val = str(df.iloc[r, col_idx]).strip() if pd.notna(df.iloc[r, col_idx]) else ''
                        if cell_val and cell_val.lower() not in ['nan', 'none', '']:
                            cell_marks.append({'row': r, 'col': col_idx, 'value': cell_val, 'type': 'l6_region', 'target': f'L6.{field_name}'})
        
        # Mark KP region on the heatmap
        kp_cfg = self._kp_region_config
        kp_start = l6_end if l6_end >= 0 else self._find_region_row(df, kp_cfg.get('region_start_keywords', 'Keyparts,KP'))
        kp_end = self._find_region_row(df, kp_cfg.get('region_end_keywords', 'Warranty,Total'), start_row=kp_start + 1 if kp_start >= 0 else 0)
        
        if kp_start >= 0:
            end_row = kp_end if kp_end > kp_start else len(df)
            # Mark start keyword row
            for c in range(min(10, df.shape[1])):
                cell_val = str(df.iloc[kp_start, c]).strip() if pd.notna(df.iloc[kp_start, c]) else ''
                if cell_val:
                    cell_marks.append({'row': kp_start, 'col': c, 'value': cell_val, 'type': 'kp_region', 'target': 'KP Start'})
            # Mark data rows
            kp_field_map = kp_cfg.get('field_mapping', {})
            if isinstance(kp_field_map, str):
                try: kp_field_map = json.loads(kp_field_map)
                except: kp_field_map = {}
            for r in range(kp_start + 1, end_row):
                for field_name, col_letter in kp_field_map.items():
                    col_idx = self._col_letter_to_index(col_letter)
                    if col_idx < df.shape[1]:
                        cell_val = str(df.iloc[r, col_idx]).strip() if pd.notna(df.iloc[r, col_idx]) else ''
                        if cell_val and cell_val.lower() not in ['nan', 'none', '']:
                            cell_marks.append({'row': r, 'col': col_idx, 'value': cell_val, 'type': 'kp_region', 'target': f'KP.{field_name}'})
        
        # Mark KP category keywords on the heatmap
        if kp_mappings:
            for mapping in kp_mappings:
                # Support both database format (keyword/category) and frontend format (keywords/variable)
                keyword = mapping.get('keyword') or mapping.get('keywords')
                category = mapping.get('category') or mapping.get('variable')
                
                if not keyword or not category:
                    continue
                
                # Handle keyword as single string or list
                keywords_list = keyword if isinstance(keyword, list) else [keyword]
                
                for kw in keywords_list:
                    kw_lower = kw.lower()
                    for r in range(len(df)):
                        for c in range(df.shape[1]):
                            cell_val = str(df.iloc[r, c]).strip() if pd.notna(df.iloc[r, c]) else ''
                            if cell_val and kw_lower in cell_val.lower():
                                # Only mark if not already marked
                                if not any(m.get('row') == r and m.get('col') == c for m in cell_marks):
                                    cell_marks.append({
                                        'row': r, 'col': c,
                                        'value': cell_val,
                                        'type': 'kp_category',
                                        'target': category
                                    })
        
        # Build grid: list of rows, each row is list of cell values
        rows = len(df) if max_row is None else min(max_row, len(df))
        cols = df.shape[1] if max_col is None else min(max_col, df.shape[1])
        grid = []
        for r in range(rows):
            row_data = []
            for c in range(cols):
                val = df.iloc[r, c]
                row_data.append(str(val).strip() if pd.notna(val) else '')
            grid.append(row_data)
        
        return {
            'grid': grid,
            'max_row': rows,
            'max_col': cols,
            'cell_marks': cell_marks,
            'meta': meta
        }

    def _extract_meta(self, df: pd.DataFrame, return_cell_marks: bool = False) -> dict:
        """Extract opportunity metadata from Excel sheet header area.
        
        Scans the first 10 rows for known keywords and extracts values.
        Also extracts L6 matching dimensions (chassis, drive_bays, psu, motherboard).
        
        Args:
            df: DataFrame from Excel sheet
            return_cell_marks: If True, return (meta, cell_marks) tuple for heatmap preview
        
        Returns:
            dict: Extracted metadata
            tuple: (meta, cell_marks) if return_cell_marks=True
        """
        meta = {}
        cell_marks = [] if return_cell_marks else None
        
        # Helper: find keyword in first N rows, return value to its right (scan entire row for first non-empty value after keyword)
        def find_keyword_value(keyword: str, max_rows: int = 10, mark_type: str = 'meta') -> str:
            keyword_lower = keyword.lower()
            for r in range(min(max_rows, len(df))):
                for c in range(min(10, df.shape[1])):
                    cell_val = str(df.iloc[r, c]).strip() if pd.notna(df.iloc[r, c]) else ''
                    if keyword_lower in cell_val.lower():
                        if return_cell_marks:
                            cell_marks.append({
                                'row': r, 'col': c,
                                'value': cell_val,
                                'type': mark_type,
                                'target': keyword
                            })
                        # Scan from c+1 to end of row for first non-empty value
                        for next_c in range(c + 1, min(10, df.shape[1])):
                            val = df.iloc[r, next_c]
                            if pd.notna(val):
                                extracted = str(val).strip()
                                if extracted and extracted.lower() not in ['', 'nan', 'none']:
                                    if return_cell_marks:
                                        cell_marks.append({
                                            'row': r, 'col': next_c,
                                            'value': extracted,
                                            'type': 'extracted',
                                            'target': keyword
                                        })
                                    return extracted
            return ''
        
        # Extract opportunity metadata from header (try multiple keyword variants)
        opportunity_name = find_keyword_value('Project Name') or find_keyword_value('商机名称')
        if opportunity_name:
            meta['opportunity_name'] = opportunity_name
        
        model_info = find_keyword_value('Model') or find_keyword_value('型号')
        if model_info:
            m = re.search(r'\((\d+)', model_info)
            if m:
                meta['model_qty'] = m.group(1)
                meta['model_name'] = model_info.split('(')[0].strip()
            else:
                meta['model_name'] = model_info
        
        fae = find_keyword_value('FAE')
        if fae and fae not in ['/', '']:
            meta['fae'] = fae
        
        # Try multiple keyword variants for l6_desc
        l6_desc = find_keyword_value('L6 Description') or find_keyword_value('PRODUCT SPEC') or find_keyword_value('产品规格')
        if l6_desc and len(l6_desc) > 10:
            meta['l6_desc'] = l6_desc
        
        date_val = find_keyword_value('Date') or find_keyword_value('日期')
        if date_val:
            meta['date'] = date_val
        
        # Fallback: try hardcoded positions if nothing found
        if 'opportunity_name' not in meta:
            try:
                val = df.iloc[1, 3]
                if pd.notna(val):
                    meta['opportunity_name'] = str(val).strip()
            except:
                pass
        if 'model_name' not in meta:
            try:
                val = df.iloc[2, 3]
                if pd.notna(val):
                    raw = str(val).strip()
                    m = re.search(r'\((\d+)', raw)
                    if m:
                        meta['model_qty'] = m.group(1)
                        meta['model_name'] = raw.split('(')[0].strip()
                    else:
                        meta['model_name'] = raw
            except:
                pass
        
        if return_cell_marks:
            return meta, cell_marks
        return meta

    def _find_region_row(self, df: pd.DataFrame, keywords_str: str, start_row: int = 0) -> int:
        """Find the first row containing any of the given keywords.
        Three-pass matching: (1) word boundary, (2) substring, (3) fuzzy match (edit distance ≤2).
        Returns row index or -1.
        """
        import re
        keywords = [k.strip() for k in keywords_str.split(',') if k.strip()]
        
        # First pass: word boundary matching
        for r in range(start_row, len(df)):
            for c in range(min(10, df.shape[1])):
                cell_val = str(df.iloc[r, c]).strip() if pd.notna(df.iloc[r, c]) else ''
                if not cell_val:
                    continue
                cell_lower = cell_val.lower()
                for kw in keywords:
                    kw_lower = kw.lower()
                    pattern = r'\b' + re.escape(kw_lower) + r'\b'
                    if re.search(pattern, cell_lower):
                        return r
        
        # Second pass: substring matching
        for r in range(start_row, len(df)):
            for c in range(min(10, df.shape[1])):
                cell_val = str(df.iloc[r, c]).strip() if pd.notna(df.iloc[r, c]) else ''
                if not cell_val:
                    continue
                cell_lower = cell_val.lower()
                for kw in keywords:
                    kw_lower = kw.lower()
                    if kw_lower in cell_lower:
                        return r
        
        # Third pass: fuzzy matching (edit distance ≤2 for typo tolerance)
        # Only apply to keywords with ≥4 chars to avoid false positives on short keywords
        def _edit_distance(s1: str, s2: str) -> int:
            """Compute Levenshtein edit distance between two strings."""
            if len(s1) < len(s2):
                return _edit_distance(s2, s1)
            if len(s2) == 0:
                return len(s1)
            prev_row = list(range(len(s2) + 1))
            for i, c1 in enumerate(s1):
                curr_row = [i + 1]
                for j, c2 in enumerate(s2):
                    insertions = prev_row[j + 1] + 1
                    deletions = curr_row[j] + 1
                    substitutions = prev_row[j] + (c1 != c2)
                    curr_row.append(min(insertions, deletions, substitutions))
                prev_row = curr_row
            return prev_row[-1]
        
        for r in range(start_row, len(df)):
            for c in range(min(10, df.shape[1])):
                cell_val = str(df.iloc[r, c]).strip() if pd.notna(df.iloc[r, c]) else ''
                if not cell_val:
                    continue
                cell_lower = cell_val.lower()
                for kw in keywords:
                    kw_lower = kw.lower()
                    # Skip fuzzy matching for short keywords (≤3 chars) to avoid false positives
                    if len(kw_lower) <= 3:
                        continue
                    # Check if any word in the cell is within edit distance 2 of the keyword
                    cell_words = re.findall(r'[a-z]+', cell_lower)
                    for word in cell_words:
                        if abs(len(word) - len(kw_lower)) <= 2:  # Quick length filter
                            if _edit_distance(word, kw_lower) <= 2:
                                return r
        
        return -1
    
    def _col_letter_to_index(self, letter: str) -> int:
        """Convert column letter (A=0, B=1, ..., Z=25, AA=26) to 0-based index."""
        letter = letter.strip().upper()
        result = 0
        for ch in letter:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    def _parse_items(self, df: pd.DataFrame) -> pd.DataFrame:
        """Parse BOM items from Excel sheet using configured L6 and KP region configs.
        
        Uses region_start_keywords and region_end_keywords to identify L6 and KP sections,
        then extracts items using field_mapping column letters.
        """
        items = []
        
        # Parse L6 region config
        l6_cfg = self._l6_region_config
        l6_field_map = l6_cfg.get('field_mapping', {})
        if isinstance(l6_field_map, str):
            try:
                l6_field_map = json.loads(l6_field_map)
            except:
                l6_field_map = {"catalogue": "D", "description": "E", "quantity": "F"}
        
        # Parse KP region config
        kp_cfg = self._kp_region_config
        kp_field_map = kp_cfg.get('field_mapping', {})
        if isinstance(kp_field_map, str):
            try:
                kp_field_map = json.loads(kp_field_map)
            except:
                kp_field_map = {"catalogue": "D", "model": "E", "quantity": "F", "price": "G"}
        
        # Find L6 region
        l6_start = self._find_region_row(df, l6_cfg.get('region_start_keywords', 'L6'))
        l6_end = self._find_region_row(df, l6_cfg.get('region_end_keywords', 'Keyparts,KP'), start_row=l6_start + 1 if l6_start >= 0 else 0)
        
        # Find KP region
        kp_start = l6_end if l6_end >= 0 else self._find_region_row(df, kp_cfg.get('region_start_keywords', 'Keyparts,KP'))
        kp_end = self._find_region_row(df, kp_cfg.get('region_end_keywords', 'Warranty,Total'), start_row=kp_start + 1 if kp_start >= 0 else 0)
        
        # Extract L6 items
        if l6_start >= 0:
            end_row = l6_end if l6_end > l6_start else len(df)
            for r in range(l6_start + 1, end_row):  # skip header row
                try:
                    cat_col = self._col_letter_to_index(l6_field_map.get('catalogue', 'D'))
                    desc_col = self._col_letter_to_index(l6_field_map.get('description', 'E'))
                    qty_col = self._col_letter_to_index(l6_field_map.get('quantity', 'F'))
                    
                    catalogue = str(df.iloc[r, cat_col]).strip() if cat_col < df.shape[1] and pd.notna(df.iloc[r, cat_col]) else ''
                    description = str(df.iloc[r, desc_col]).strip() if desc_col < df.shape[1] and pd.notna(df.iloc[r, desc_col]) else ''
                    qty = 1
                    if qty_col < df.shape[1] and pd.notna(df.iloc[r, qty_col]):
                        try:
                            qty = int(float(df.iloc[r, qty_col]))
                        except:
                            qty = 1
                    
                    if not catalogue or catalogue.lower() in ['nan', 'none', '', 'catalogue', 'component description']:
                        continue
                    
                    items.append({
                        'category': 'L6',
                        'part_name': catalogue,
                        'spec': description,
                        'qty': qty,
                        'currency': 'RMB'
                    })
                except:
                    continue
        
        # Extract KP items
        if kp_start >= 0:
            end_row = kp_end if kp_end > kp_start else len(df)
            for r in range(kp_start + 1, end_row):  # skip header row
                try:
                    cat_col = self._col_letter_to_index(kp_field_map.get('catalogue', 'D'))
                    model_col = self._col_letter_to_index(kp_field_map.get('model', 'E'))
                    qty_col = self._col_letter_to_index(kp_field_map.get('quantity', 'F'))
                    price_col = self._col_letter_to_index(kp_field_map.get('price', 'G'))
                    
                    catalogue = str(df.iloc[r, cat_col]).strip() if cat_col < df.shape[1] and pd.notna(df.iloc[r, cat_col]) else ''
                    model = str(df.iloc[r, model_col]).strip() if model_col < df.shape[1] and pd.notna(df.iloc[r, model_col]) else ''
                    qty = 1
                    if qty_col < df.shape[1] and pd.notna(df.iloc[r, qty_col]):
                        try:
                            qty = int(float(df.iloc[r, qty_col]))
                        except:
                            qty = 1
                    
                    price = None
                    if price_col < df.shape[1] and pd.notna(df.iloc[r, price_col]):
                        price_val = df.iloc[r, price_col]
                        if isinstance(price_val, str) and price_val.startswith('='):
                            try:
                                formula = price_val[1:]
                                price = _safe_eval_math(formula)
                            except:
                                pass
                        else:
                            try:
                                price = float(price_val)
                            except:
                                pass
                    
                    if not catalogue or catalogue.lower() in ['nan', 'none', '', 'catalogue']:
                        continue
                    
                    # Determine if USD (e.g., CPU items)
                    is_usd = False
                    if 'cpu' in catalogue.lower() or 'processor' in catalogue.lower():
                        if 'usd' in model.lower() or '$' in model:
                            is_usd = True
                    
                    items.append({
                        'category': 'Key Parts',
                        'part_name': catalogue,
                        'spec': model,
                        'qty': qty,
                        'currency': 'USD' if is_usd else 'RMB'
                    })
                except:
                    continue
        
        # Extract Warranty items (if Warranty region exists)
        warranty_start = self._find_region_row(df, 'Warranty', start_row=kp_start + 1 if kp_start >= 0 else 0)
        warranty_end = self._find_region_row(df, 'Total,Total Price', start_row=warranty_start + 1 if warranty_start >= 0 else 0)
        
        if warranty_start >= 0:
            end_row = warranty_end if warranty_end > warranty_start else len(df)
            for r in range(warranty_start + 1, end_row):
                try:
                    # Warranty format: row number, type (L6/KP), description
                    type_col = 2  # Column C (0-indexed)
                    desc_col = 3  # Column D
                    
                    warranty_type = str(df.iloc[r, type_col]).strip() if pd.notna(df.iloc[r, type_col]) else ''
                    description = str(df.iloc[r, desc_col]).strip() if pd.notna(df.iloc[r, desc_col]) else ''
                    
                    if not description or description.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Extract warranty years from description — let user fill in manually
                    years = None
                    
                    items.append({
                        'category': 'Warranty',
                        'part_name': warranty_type,
                        'spec': description,
                        'qty': 1,
                        'currency': 'RMB',
                        'warranty_years': years
                    })
                except:
                    continue
        
        return pd.DataFrame(items) if items else pd.DataFrame()

    # ==================== 2. Price Enrichment (via Repository) ====================

    def enrich_config(self, items_df: pd.DataFrame, meta: Optional[dict] = None) -> pd.DataFrame:
        """Enrich items with DB price match status (NO auto-fill).
        Uses kp_repo instead of direct sqlite3 queries."""
        if items_df.empty:
            return items_df

        # Fetch all latest KP prices at once (one query via repo)
        kp_latest = self.kp_repo.get_latest_prices()
        kp_dict = {r['model'].lower().strip(): r['price'] for r in kp_latest}

        items = items_df.copy()
        items['match_status'] = ""
        items['db_price'] = None
        items['base_price'] = items.get('price', 0)
        items['is_usd_cpu'] = items.get('currency', 'RMB') == 'USD'
        items['profit_margin'] = 10.0  # default

        for idx, row in items.iterrows():
            cat = row['category']
            name = str(row['part_name']).lower().strip()
            spec = str(row.get('spec', '')).lower().strip()
            original_price = row.get('price')
            db_price = None

            if cat == 'Key Parts':
                if name in kp_dict:
                    db_price = kp_dict[name]
                elif spec and len(spec) > 2:
                    fuzzy = self.kp_repo.fuzzy_match_price(spec)
                    if fuzzy:
                        db_price = fuzzy['price']

            items.at[idx, 'db_price'] = db_price

            if db_price is not None:
                if pd.isna(original_price) or original_price == 0:
                    items.at[idx, 'match_status'] = f"⚠️ 待填入 [DB={db_price}]"
                else:
                    if abs(float(original_price) - db_price) > self._price_diff_threshold:
                        items.at[idx, 'match_status'] = f"⚠️ 差异 (Excel: {original_price}, DB: {db_price})"
                    else:
                        items.at[idx, 'match_status'] = f"✅ 一致 [DB={db_price}]"
            else:
                if pd.isna(original_price) or original_price == 0:
                    items.at[idx, 'match_status'] = "❌ 缺失 (请填写)"
                else:
                    items.at[idx, 'match_status'] = "🆕 新部件"

        # NaN sanitization for JSON serialization
        for col in items.columns:
            if items[col].dtype == object:
                items[col] = items[col].fillna("")
            else:
                items[col] = items[col].fillna(0)

        return items

    # ==================== 3. KP Sync & History (via Repository) ====================

    def get_kp_price_history(self, model: str, limit: int = 10) -> list:
        return self.kp_repo.get_price_history(model, limit)

    def sync_kp_prices_to_db(self, configs_data: dict) -> int:
        """Compare and insert new KP prices into DB if different from latest."""
        today = datetime.now().strftime('%Y-%m-%d')
        new_records = 0

        # Use configurable category mapping
        cat_map = self._kp_cat_map

        # Batch-load all latest prices once (avoid N+1 queries)
        latest_prices = {}
        try:
            all_latest = self.kp_repo.get_latest_prices()
            latest_prices = {
                r['model'].lower().strip(): float(r['price'])
                for r in all_latest
                if r.get('model') and r.get('price') is not None
            }
        except Exception:
            pass

        pending_inserts = []  # Collect for batch insert

        for cfg_name, items_df in configs_data.items():
            if items_df.empty:
                continue
            if 'category' not in items_df.columns:
                continue

            kp_items = items_df[items_df['category'] == 'Key Parts'].copy()
            if kp_items.empty:
                continue
            for idx, row in kp_items.iterrows():
                part_name = str(row.get('part_name', '')).strip()
                spec = str(row.get('spec', '')).strip()

                category = None
                model = None
                lower_name = part_name.lower()
                for keyword, std_cat in cat_map.items():
                    if keyword in lower_name:
                        category = std_cat
                        model = spec if spec else part_name
                        break
                if category is None:
                    category = row.get('category', 'Key Parts')
                    model = part_name
                if not model:
                    continue

                new_price = row.get('base_price', 0)
                if pd.isna(new_price) or new_price == 0:
                    continue

                # In-memory lookup instead of DB query per item
                db_price = latest_prices.get(model.lower().strip())
                if db_price is not None and abs(float(new_price) - db_price) < 0.01:
                    continue

                pending_inserts.append((category, model, round(float(new_price), 2)))

        # Batch insert all new prices
        for category, model, price in pending_inserts:
            self.kp_repo.insert_price(category, model, price, 'RMB', today, '报价系统更新')
            new_records += 1

        return new_records

    # ==================== 4. Project CRUD (via Repository) ====================

    def save_opportunity(self, opportunity_info: dict, configs_data: dict, config_descriptions: dict = None, config_quantities: dict = None, config_server_models: dict = None, config_warranty_info: dict = None) -> dict:
        """Save opportunity meta + items to projects.db + quotations.db."""
        import time, random
        opportunity_id = opportunity_info.get('opportunity_id', '').strip()
        if not opportunity_id:
            # Auto-generate: {opportunity_name}_{timestamp}_{random4}
            name = opportunity_info.get('opportunity_name', 'opportunity')
            ts = time.strftime('%m%d%H%M%S')
            rand = f'{random.randint(0, 0xFFFF):04x}'
            opportunity_id = f'{name}_{ts}_{rand}'
            opportunity_info['opportunity_id'] = opportunity_id
        self.opportunity_repo.create_or_update_opportunity(opportunity_id, opportunity_info)
        
        # Convert configs_data to items and save via QuotationRepository
        q_repo = self._get_quotation_repo()
        try:
            # Create a quotation for this opportunity
            quotation = q_repo.create(
                opportunity_id=opportunity_id,
                quotation_date=datetime.now().strftime('%Y-%m-%d')
            )
            quotation_id = quotation.quotation_id
            
            # Flatten all config items into a single list
            all_items = []
            for cfg_name, cfg_data in configs_data.items():
                if isinstance(cfg_data, pd.DataFrame):
                    items = cfg_data.to_dict('records')
                elif isinstance(cfg_data, dict):
                    items = [cfg_data]
                elif isinstance(cfg_data, list):
                    items = cfg_data
                else:
                    continue
                for item in items:
                    item['config_name'] = item.get('config_name', cfg_name)
                    # Ensure required fields
                    item.setdefault('base_price', 0.0)
                    item.setdefault('final_price', 0.0)
                    item.setdefault('profit_margin', 0.0)
                    item.setdefault('spec', '')
                    item.setdefault('category', '')
                    item.setdefault('part_name', '')
                    item.setdefault('qty', 0)
                    item.setdefault('is_usd_cpu', False)
                all_items.extend(items)
            
            # Calculate totals before saving
            l6_total = sum(i.get('final_price', 0) for i in all_items if i.get('category') == 'L6')
            kp_total = sum(i.get('final_price', 0) for i in all_items if i.get('category') == 'Key Parts')
            grand_total = l6_total + kp_total
            total_qty = sum(i.get('qty', 0) for i in all_items if i.get('category') == 'L6')
            config_count = len(set(i.get('config_name', 'CFG1') for i in all_items))
            
            # Update quotation with computed totals
            q_repo.update(quotation_id,
                l6_price=l6_total,
                total_price=grand_total,
                total_qty=total_qty,
                config_count=config_count
            )
            
            # Save config descriptions if provided
            if config_descriptions:
                q_repo.update(quotation_id, config_descriptions=config_descriptions)
            
            # Save config quantities if provided
            if config_quantities:
                q_repo.update(quotation_id, config_quantities=config_quantities)
            
            # Save config server_models if provided
            if config_server_models:
                q_repo.update(quotation_id, config_server_models=config_server_models)
            
            # Save config warranty_info if provided
            if config_warranty_info:
                q_repo.update(quotation_id, config_warranty_info=config_warranty_info)
            
            # Save items
            item_count = q_repo.save_items(quotation_id, all_items)
        finally:
            q_repo.close()
        
        return {"status": "success", "items_saved": item_count, "opportunity_id": opportunity_id, "quotation_id": quotation_id}

    def get_opportunity_details(self, opportunity_id: str) -> Optional[dict]:
        """Get opportunity with its latest quotation's items (for export)."""
        q_repo = self._get_quotation_repo()
        try:
            project_dict = self.opportunity_repo.get_opportunity(opportunity_id)

            if not project_dict:
                return None
            
            # Get all active quotations for this opportunity
            quotations = q_repo.get_by_opportunity(opportunity_id)
            
            # Start with all fields from project_dict (includes expanded extra_fields)
            meta = dict(project_dict)
            meta['date'] = ''
            
            # Enrich meta from the latest quotation
            if quotations:
                latest = quotations[0]  # already ordered by version desc
                meta['date'] = latest.quotation_date or meta.get('date', '')
                meta['quotation_id'] = latest.quotation_id
                meta['version'] = latest.version
                # Pass config_quantities for cover sheet rendering
                if hasattr(latest, 'config_quantities') and latest.config_quantities:
                    meta['config_quantities'] = latest.config_quantities
                # Pass config_descriptions for config sheet rendering
                if hasattr(latest, 'config_descriptions') and latest.config_descriptions:
                    meta['config_descriptions'] = latest.config_descriptions
                # Pass config_server_models for config sheet rendering
                if hasattr(latest, 'config_server_models') and latest.config_server_models:
                    meta['config_server_models'] = latest.config_server_models
                # Pass config_warranty_info for per-config warranty restoration
                if hasattr(latest, 'config_warranty_info') and latest.config_warranty_info:
                    meta['config_warranty_info'] = latest.config_warranty_info
            
            # Build configs from quotations
            configs = {}
            for quo in quotations:
                items = q_repo.get_items(quo.quotation_id)
                cfg_items = [item.to_dict() for item in items]
                # Group by config_name
                for item in cfg_items:
                    cfg_name = item.get('config_name', 'CFG1')
                    if cfg_name not in configs:
                        configs[cfg_name] = []
                    configs[cfg_name].append(item)
            
            # Fallback: if no quotations, try to return empty configs
            if not configs:
                configs['CFG1'] = []
            
            return {'meta': meta, 'configs': configs, 'quotations': [q.to_dict() for q in quotations]}
        finally:
            q_repo.close()

    def update_project_meta(self, opportunity_id: str, updates: dict) -> bool:
        return self.opportunity_repo.update_meta(opportunity_id, updates)

    # ==================== 5. Excel Export (template-driven) ====================

    def _load_config(self) -> dict:
        """Load config from system_config DB table (single source of truth)."""
        from app.repository.system_config_repo import SystemConfigRepository
        repo = SystemConfigRepository()
        try:
            return {
                "tax_rate": repo.get_value("tax_rate", 0.13),
                "usd_to_rmb": repo.get_value("usd_to_rmb", 7.0),
                "profit_margin": repo.get_value("profit_margin", 0.1),
                "warranty_fee_rate": repo.get_value("warranty_fee_rate", 0.02),
            }
        finally:
            repo.close()

    def _resolve_font(self, font_dict: dict) -> Font:
        """Convert font dict from template JSON to openpyxl Font."""
        return Font(
            name=font_dict.get('name', '宋体'),
            size=font_dict.get('size', 11),
            bold=font_dict.get('bold', False),
            italic=font_dict.get('italic', False),
            color=font_dict.get('color', '000000')
        )


