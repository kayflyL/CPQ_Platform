"""
Excel → Univer workbook_snapshot 转换服务

职责：
- excel_to_snapshot: 上传 Excel → 转为 Univer 格式
"""
import io
import openpyxl
from typing import Tuple


def excel_to_snapshot(file_stream: io.BytesIO) -> Tuple[dict, dict]:
    """
    将 Excel 文件转为 Univer workbook_snapshot 格式
    
    Returns:
        (workbook_snapshot, sheet_config)
        
    workbook_snapshot 结构:
    {
        "sheetOrder": ["sheet-1", "sheet-2"],
        "sheets": {
            "sheet-1": {
                "id": "sheet-1",
                "name": "Sheet1",
                "cellData": { "0": { "0": { "v": "value", "s": {...} } } },
                "rowData": { "0": { "h": 20 } },
                "columnData": { "0": { "w": 100 } },
                "mergeData": [{ "startRow": 0, "startColumn": 0, "endRow": 1, "endColumn": 1 }],
                "rowCount": 100,
                "columnCount": 26
            }
        }
    }
    """
    # 使用 data_only=False 保留公式
    wb = openpyxl.load_workbook(file_stream, data_only=False)
    
    sheets = {}
    sheet_order = []
    
    for idx, ws in enumerate(wb.worksheets):
        sheet_id = f"sheet-{idx + 1}"
        sheet_order.append(sheet_id)
        
        # 转换单元格数据
        cell_data = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    row_idx = str(cell.row - 1)  # 0-indexed
                    col_idx = str(cell.column - 1)  # 0-indexed
                    
                    if row_idx not in cell_data:
                        cell_data[row_idx] = {}
                    
                    # 优先使用公式，如果没有公式则使用缓存值
                    cell_obj = {}
                    if cell.data_type == 'f':  # 公式类型
                        cell_obj["f"] = cell.value  # 保留公式
                        # 如果有缓存值，也保存一份用于预览
                        if cell._value is not None:
                            cell_obj["v"] = cell._value
                    else:
                        cell_obj["v"] = cell.value
                    
                    # 转换样式
                    style = _convert_cell_style(cell)
                    if style:
                        cell_obj["s"] = style
                    
                    cell_data[row_idx][col_idx] = cell_obj
        
        # 转换合并单元格
        merge_data = []
        for merge_range in ws.merged_cells.ranges:
            merge_data.append({
                "startRow": merge_range.min_row - 1,
                "startColumn": merge_range.min_col - 1,
                "endRow": merge_range.max_row - 1,
                "endColumn": merge_range.max_col - 1,
            })
        
        # 转换行高（Excel磅 → Univer像素，1磅 = 4/3像素）
        row_data = {}
        for row_idx, dim in ws.row_dimensions.items():
            if dim.height:
                row_data[str(row_idx - 1)] = {"h": dim.height * 4 / 3}
        
        # 转换列宽（Excel字符数 → Univer像素，1字符 ≈ 8.66像素）
        column_data = {}
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                column_data[str(col_idx)] = {"w": dim.width * 8.66}
        
        sheets[sheet_id] = {
            "id": sheet_id,
            "name": ws.title,
            "cellData": cell_data,
            "rowData": row_data,
            "columnData": column_data,
            "mergeData": merge_data,
            "rowCount": max(ws.max_row or 1, 1000),
            "columnCount": max(ws.max_column or 1, 26),
        }
    
    wb.close()
    
    workbook_snapshot = {
        "sheetOrder": sheet_order,
        "sheets": sheets,
    }
    
    # 默认 sheet_config: 第一个为 cover，其他为 config
    sheet_config = {
        "cover": {"sheetId": sheet_order[0]} if sheet_order else None,
        "config": {
            "sheetId": sheet_order[1] if len(sheet_order) > 1 else None,
            "nameTemplate": "{cfg_name}",
            "splitByConfig": True,
        } if len(sheet_order) > 1 else None,
    }
    
    return workbook_snapshot, sheet_config


def _convert_cell_style(cell) -> dict:
    """转换单元格样式为 Univer 格式"""
    style = {}
    
    # 字体
    if cell.font:
        font = cell.font
        style["ff"] = font.name or "Arial"  # fontFamily
        style["fs"] = font.size or 11  # fontSize
        if font.bold:
            style["bl"] = 1  # bold
        if font.italic:
            style["it"] = 1  # italic
        if font.color and font.color.rgb:
            rgb_str = str(font.color.rgb)
            # 只处理有效的 hex RGB 值，忽略 ThemeColor 等复杂类型
            if len(rgb_str) >= 6 and all(c in "0123456789ABCDEFabcdef" for c in rgb_str[-6:]):
                style["fc"] = rgb_str[2:] if len(rgb_str) >= 8 else rgb_str[-6:]  # fontColor (去掉 alpha)
    
    # 背景色
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb_str = str(cell.fill.fgColor.rgb)
        # 只处理有效的 hex RGB 值，忽略 ThemeColor 等复杂类型
        if len(rgb_str) >= 6 and all(c in "0123456789ABCDEFabcdef" for c in rgb_str[-6:]):
            style["bg"] = rgb_str[2:] if len(rgb_str) >= 8 else rgb_str[-6:]  # backgroundColor
    
    # 对齐
    if cell.alignment:
        align = cell.alignment
        if align.horizontal:
            style["ht"] = _map_horizontal_align(align.horizontal)
        if align.vertical:
            style["vt"] = _map_vertical_align(align.vertical)
        if align.wrap_text:
            style["tb"] = 2  # textBreak (wrap)
    
    # 边框（简化处理）
    if cell.border:
        border = cell.border
        if any([border.top, border.bottom, border.left, border.right]):
            style["bd"] = _convert_border(border)
    
    return style if style else None


def _map_horizontal_align(align: str) -> int:
    """映射水平对齐"""
    mapping = {
        "left": 1,
        "center": 2,
        "right": 3,
        "fill": 4,
        "justify": 5,
    }
    return mapping.get(align, 0)


def _map_vertical_align(align: str) -> int:
    """映射垂直对齐"""
    mapping = {
        "top": 1,
        "middle": 2,
        "bottom": 3,
    }
    return mapping.get(align, 0)




def _convert_border(border) -> dict:
    """转换边框"""
    bd = {}
    if border.top and border.top.style:
        bd["t"] = {"s": 1, "cl": {"rgb": "000000"}}
    if border.bottom and border.bottom.style:
        bd["b"] = {"s": 1, "cl": {"rgb": "000000"}}
    if border.left and border.left.style:
        bd["l"] = {"s": 1, "cl": {"rgb": "000000"}}
    if border.right and border.right.style:
        bd["r"] = {"s": 1, "cl": {"rgb": "000000"}}
    return bd


