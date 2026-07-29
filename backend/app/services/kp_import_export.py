"""KP 配件批量导入 / 导出服务。

与报价工作台的 excel_parser(规则引擎)定位不同:本模块面向字段固定的结构化主数据,
采用「固定模板 + 死列名 + 轻量中英别名」的解析方式,不做用户可配置规则。

三类入口:
  - ExportService.export_workbook  全量配件 → xlsx bytes
  - TemplateService.build_template  下载导入模板 → xlsx bytes
  - ImportService.parse/classify/commit  上传 → 预览 → 写入
"""
import io
from datetime import date
from typing import Any, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# 固定列:(canonical key, 中文表头) —— 顺序即导出/模板顺序
FIXED_COLUMNS = [
    ("category_name", "分类"),
    ("name",          "名称"),
    ("oem_sku",       "料号"),
    ("alt_sku",       "副料号"),
    ("brand",         "品牌"),
    ("short_desc",    "简述"),
    ("condition",     "成色"),
    ("lead_time",     "货期"),
    ("moq",           "起订量"),
    ("price",         "价格"),
    ("currency",      "币种"),
    ("price_date",    "价格日期"),
    ("price_note",    "价格备注"),
]

# 中英别名 → canonical key(大小写/空格不敏感)
ALIASES = {
    "category_name": ["分类", "类别", "category", "cat", "分类名称", "key part", "key_part"],
    "name":          ["名称", "配件名称", "型号", "name", "model", "配件名", "品名"],
    "oem_sku":       ["料号", "oem_sku", "oem", "sku", "pn", "part_number", "原厂料号", "mpn"],
    "alt_sku":       ["副料号", "alt_sku", "alt", "替代料号", "替代型号"],
    "brand":         ["品牌", "brand", "厂商", "厂家", "manufacturer"],
    "short_desc":    ["简述", "short_desc", "描述", "说明", "short description"],
    "condition":     ["成色", "condition", "状态", "新旧"],
    "lead_time":     ["货期", "lead_time", "交期", "交货期"],
    "moq":           ["起订量", "moq", "最小起订", "minimum order"],
    "price":         ["价格", "price", "单价", "报价", "金额"],
    "currency":      ["币种", "currency", "货币"],
    "price_date":    ["价格日期", "price_date", "报价日期", "调价日期", "报价时间"],
    "price_note":    ["备注", "价格备注", "桃红备注", "note", "价格说明"],
}

_ALIAS_LOOKUP = {
    str(a).strip().lower().replace(" ", ""): canon
    for canon, aliases in ALIASES.items() for a in aliases
}

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().lower().replace(" ", "")


def _canonical_col(raw: str) -> Optional[str]:
    """原始列名 → canonical key;spec. 开头原样返回(normalized);无法识别返回 None。"""
    n = _norm(raw)
    if not n:
        return None
    if n.startswith("spec.") or n.startswith("spec:"):
        return "spec"
    return _ALIAS_LOOKUP.get(n)


def _strip_spec_prefix(raw_col: Any) -> str:
    s = str(raw_col).strip()
    low = s.lower()
    if low.startswith("spec."):
        return s[5:].strip()
    if low.startswith("spec:"):
        return s[5:].strip()
    return s


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v).strip()
    return s if s else None


def _parse_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None
    return None if pd.isna(f) else f


def _parse_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _latest_price(part) -> Optional[Any]:
    hist = part.price_history or []
    if not hist:
        return None
    return sorted(hist, key=lambda h: (h.price_date or date.min, h.id), reverse=True)[0]


# ============================================================
# 导出
# ============================================================
class ExportService:
    @staticmethod
    def export_workbook(parts: List[Any]) -> bytes:
        """parts: KPPart ORM 列表(已 eager load specs/category/price_history)。"""
        spec_keys = sorted({s.spec_key for p in parts for s in (p.specs or []) if s.spec_key})
        buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "配件"

        header = [label for _, label in FIXED_COLUMNS] + [f"spec.{k}" for k in spec_keys]
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for p in parts:
            latest = _latest_price(p)
            row = [
                p.category.name if p.category else "",
                p.name or "",
                p.oem_sku or "",
                p.alt_sku or "",
                p.brand or "",
                p.short_desc or "",
                p.condition or "",
                p.lead_time or "",
                p.moq if p.moq is not None else "",
                latest.price if latest else "",
                latest.currency if latest else "RMB",
                latest.price_date.isoformat() if (latest and latest.price_date) else "",
            ]
            spec_map = {s.spec_key: (s.spec_value or "") for s in (p.specs or [])}
            row.extend(spec_map.get(k, "") for k in spec_keys)
            ws.append(row)

        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def media_type() -> str:
        return _XLSX_MEDIA


# ============================================================
# 导入模板
# ============================================================
class TemplateService:
    @staticmethod
    def build_template(spec_keys: List[str]) -> bytes:
        buf = io.BytesIO()
        wb = Workbook()

        ws = wb.active
        ws.title = "配件"
        header = [label for _, label in FIXED_COLUMNS] + [f"spec.{k}" for k in spec_keys]
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.column_dimensions["B"].width = 28  # 名称列宽一点

        ws2 = wb.create_sheet("说明")
        lines = [
            "KP 配件批量导入说明",
            "",
            "1. 必填列:名称。其余列可留空。",
            "2. 去重规则:优先按「料号」判断是否已存在;料号为空则按「名称」。",
            "   命中 0 条 → 新增;命中 1 条 → 更新;命中多条 → 标记冲突并跳过。",
            "3. 规格列:任何 spec. 开头的列都会写进配件规格(键 = 列名去掉 spec. 前缀)。",
            "   可自行追加 spec.任意键 的列,不必局限于模板预置的规格。",
            "4. 价格:填了「价格」会追加一条价格历史(不覆盖),可附「价格日期」。",
            "5. 列名容错:中英文别名均接受(料号 / oem_sku / sku / pn 等效),大小写与空格不敏感。",
            "6. 分类:填文本名称,不存在自动新建;留空归入「未分类」。",
        ]
        for ln in lines:
            ws2.append([ln])
        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 90

        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def media_type() -> str:
        return _XLSX_MEDIA


# ============================================================
# 导入
# ============================================================
class ImportService:
    @staticmethod
    def parse(file_bytes: bytes) -> dict:
        """返回 {rows: [...], has_spec_columns: bool}。"""
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
        col_map: dict = {}        # canonical -> raw col(首个优先)
        spec_cols: list = []      # [(raw_col)] 仅记原始列名
        has_spec = False
        for raw in df.columns:
            canon = _canonical_col(raw)
            if canon is None:
                continue
            if canon == "spec":
                has_spec = True
                spec_cols.append(raw)
            else:
                col_map.setdefault(canon, raw)

        def gv(canon: str, row) -> Optional[str]:
            raw = col_map.get(canon)
            return _clean(row.get(raw)) if raw is not None else None

        rows = []
        for idx, df_row in df.iterrows():
            name = gv("name", df_row)
            oem = gv("oem_sku", df_row)
            brand = gv("brand", df_row)
            # 全空行跳过
            if not name and not oem and not brand:
                continue
            specs = []
            for raw in spec_cols:
                val = _clean(df_row.get(raw))
                if val is None:
                    continue
                specs.append({"key": _strip_spec_prefix(raw), "value": val})
            rows.append({
                "_row_index": int(idx) + 2,  # Excel 行号(1-based + 表头)
                "category_name": gv("category_name", df_row),
                "name": name,
                "oem_sku": oem,
                "alt_sku": gv("alt_sku", df_row),
                "brand": brand,
                "short_desc": gv("short_desc", df_row),
                "condition": gv("condition", df_row),
                "lead_time": gv("lead_time", df_row),
                "moq": _parse_int(gv("moq", df_row)),
                "price": _parse_float(gv("price", df_row)),
                "currency": gv("currency", df_row) or "RMB",
                "price_date": gv("price_date", df_row),
                "price_note": gv("price_note", df_row),
                "specs": specs,
            })
        return {"rows": rows, "has_spec_columns": has_spec}

    @staticmethod
    def classify(rows: list, repo) -> list:
        preview = []
        for r in rows:
            name = r.get("name")
            oem = r.get("oem_sku")
            if not name:
                preview.append({**r, "action": "invalid", "message": "名称为空,跳过"})
                continue
            existing = repo.find_parts_by_dedupe_key(oem_sku=oem, name=name)
            if len(existing) == 0:
                preview.append({**r, "action": "new", "message": "将新增"})
            elif len(existing) == 1:
                ex = existing[0]
                preview.append({**r, "action": "update", "existing_id": ex.id,
                                "message": f"将更新 #{ex.id}"})
            else:
                key = f"料号 {oem}" if oem else f"名称 {name}"
                preview.append({**r, "action": "conflict",
                                "message": f"{key} 命中 {len(existing)} 条,请先去重"})
        return preview

    @staticmethod
    def summarize_preview(preview: list) -> dict:
        def count(act: str) -> int:
            return sum(1 for p in preview if p.get("action") == act)
        return {
            "new": count("new"), "update": count("update"),
            "conflict": count("conflict"), "invalid": count("invalid"),
            "total": len(preview),
        }

    @staticmethod
    def commit(preview_rows: list, repo) -> dict:
        created = updated = skipped = failed = price_dup = 0
        errors = []
        session_cache = {}  # dedupe_key -> part_id:同次导入内同件多行只建一次,后续行追加价格
        for r in preview_rows:
            action = r.get("action")
            if action in ("conflict", "invalid"):
                skipped += 1
                continue
            try:
                dkey = (r.get("oem_sku") or "").strip() or f"NAME::{(r.get('name') or '').strip()}"
                # 仅把"本行实际提供值"的字段塞进 payload:缺列 = 不动已有字段(避免窄表导入清空 DB 原有值)
                payload = {"name": r["name"]}
                for fld in ("oem_sku", "alt_sku", "brand", "short_desc", "lead_time"):
                    v = r.get(fld)
                    if v is not None and v != "":
                        payload[fld] = v
                if r.get("condition"):
                    payload["condition"] = r["condition"]
                if r.get("moq") is not None:
                    payload["moq"] = r["moq"]
                if r.get("category_name"):
                    payload["category_id"] = repo.find_or_create_category_by_name(r["category_name"])
                specs_this = r.get("specs") or []
                if specs_this:
                    payload["specs"] = specs_this  # 该行有规格值才全量替换;空则保留原 specs

                cached_id = session_cache.get(dkey)
                if cached_id is not None:
                    # 本次会话已为该件建过记录,后续行仅追加价格历史
                    part_id = cached_id
                    updated += 1
                elif action == "new":
                    created_part = repo.create_part(payload)
                    part_id = created_part["id"]
                    session_cache[dkey] = part_id
                    created += 1
                else:  # update
                    repo.update_part(r["existing_id"], payload)
                    part_id = r["existing_id"]
                    session_cache[dkey] = part_id
                    updated += 1

                if r.get("price") is not None:
                    if repo.price_history_exists(part_id, r["price"], r.get("currency"), r.get("price_date")):
                        price_dup += 1
                    else:
                        repo.add_price_history(
                            part_id=part_id,
                            price=r["price"],
                            currency=r.get("currency") or "RMB",
                            price_date=r.get("price_date"),
                            note=r.get("price_note") or "",
                            source="import",
                        )
            except Exception as e:
                failed += 1
                errors.append({
                    "row": r.get("_row_index"),
                    "name": r.get("name"),
                    "message": str(e),
                })
        return {
            "created": created, "updated": updated,
            "skipped": skipped, "failed": failed,
            "price_duplicate": price_dup, "errors": errors,
        }
